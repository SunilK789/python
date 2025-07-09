import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from datetime import datetime
import re
import os
import subprocess
from openpyxl import Workbook

headers = {
    'User-Agent': 'Mozilla/5.0'
}

base_url = "https://www.screener.in/ipo/recent/?page={}"
ipo_data = []

def extract_clean_text(td):
    text = "".join(td.stripped_strings)
    return re.sub(r'[^\d.,%-]', '', text)

def format_date(date_str):
    try:
        parsed = datetime.strptime(date_str.strip(), "%d %b %Y")
        return parsed.strftime("%d-%b-%y")
    except:
        return date_str.strip()

def get_last_page():
    url = base_url.format(1)
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    # Adjust this selector if the pagination uses a different tag/class
    pagination = soup.find("div", class_="pagination")
    if not pagination:
        print("No pagination found, defaulting to 1 page")
        return 1

    page_numbers = []
    for a in pagination.find_all("a"):
        href = a.get('href', '')
        text = a.get_text(strip=True)

        if "Next" in text:
            continue

        match = re.search(r'page=(\d+)', href)
        if match:
            page_numbers.append(int(match.group(1)))

    if page_numbers:
        last_page = max(page_numbers)
        print(f"Detected last page: {last_page}")
        return last_page
    else:
        print("No page numbers found, defaulting to 1 page")
        return 1

last_page = get_last_page()

for page in range(1, last_page + 1):
    url = base_url.format(page)
    print(f"Fetching page {page} of {last_page}...")
    print(f"URL {url}")

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    table = soup.find("table", class_="data-table")
    if not table:
        print(f"No table found on page {page}")
        continue

    rows = table.find_all("tr")[1:]  # skip header
    for row in rows:
        cols = row.find_all("td")
        if len(cols) >= 6:
            name_link_tag = cols[0].find("a")
            name = name_link_tag.text.strip() if name_link_tag else cols[0].text.strip()
            relative_url = name_link_tag['href'] if name_link_tag and 'href' in name_link_tag.attrs else ''
            full_url = f"https://www.screener.in{relative_url}" if relative_url else ''

            list_date = format_date(cols[1].text.strip())
            ipo_price_str = extract_clean_text(cols[3])
            current_price_str = extract_clean_text(cols[4])

            if not ipo_price_str or not current_price_str:
                continue

            try:
                ipo_price = float(ipo_price_str.replace(',', ''))
                current_price = float(current_price_str.replace(',', ''))
            except ValueError:
                continue

            ipo_data.append({
                "Name": name,
                "Profile Link": full_url,
                "List Date": list_date,
                "IPO MCap (Cr)": extract_clean_text(cols[2]),
                "IPO Price": ipo_price_str,
                "Current Price": current_price_str,
                "Change %": extract_clean_text(cols[5]),
                "IPO Price Float": ipo_price,
                "Current Price Float": current_price
            })

    time.sleep(1)

# Prepare output paths
today_str = datetime.now().strftime("%Y-%m-%d")
base_dir = r"D:\IPO Data"
output_dir = os.path.join(base_dir, today_str)
os.makedirs(output_dir, exist_ok=True)

timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
df = pd.DataFrame(ipo_data)

def save_excel_with_hyperlinks(df_input, path, sheet_name="Sheet1"):
    df_to_export = df_input.drop(columns=["IPO Price Float", "Current Price Float"])
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(df_to_export.columns.tolist())

    for row in df_to_export.itertuples(index=False):
        excel_row = []
        for i, value in enumerate(row):
            col_name = df_to_export.columns[i]
            if col_name == "Profile Link" and value:
                excel_row.append(f'=HYPERLINK("{value}", "View")')
            else:
                excel_row.append(value)
        ws.append(excel_row)

    wb.save(path)

all_ipo_filename = os.path.join(output_dir, f"screener_ipo_data_{timestamp}.xlsx")
save_excel_with_hyperlinks(df, all_ipo_filename, "All IPOs")
print(f"Saved all IPO data to {all_ipo_filename}")

filtered_filename = os.path.join(output_dir, f"ipo_filtered_data_{timestamp}.xlsx")
wb = Workbook()
sheet_labels = [
    ("≥ 1x IPO Price", 1.0),
    (f"≥ 50% from IPO Price", 1.5),
    (f"≥ 100% from IPO Price", 2.0),
    (f"≥ 150% from IPO Price", 2.5),
]

wb.remove(wb.active)

for sheet_name, factor in sheet_labels:
    filtered_df = df[df["Current Price Float"] >= factor * df["IPO Price Float"]].copy()
    df_to_export = filtered_df.drop(columns=["IPO Price Float", "Current Price Float"])

    ws = wb.create_sheet(title=sheet_name)
    ws.append(df_to_export.columns.tolist())

    for row in df_to_export.itertuples(index=False):
        excel_row = []
        for i, value in enumerate(row):
            col_name = df_to_export.columns[i]
            if col_name == "Profile Link" and value:
                excel_row.append(f'=HYPERLINK("{value}", "View")')
            else:
                excel_row.append(value)
        ws.append(excel_row)

wb.save(filtered_filename)
print(f"Saved filtered IPO data to {filtered_filename}")

subprocess.Popen(['start', '', all_ipo_filename], shell=True)
subprocess.Popen(['start', '', filtered_filename], shell=True)
